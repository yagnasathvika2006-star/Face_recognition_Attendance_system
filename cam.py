import cv2
import face_recognition
import os
import tkinter as tk
from tkinter import simpledialog, messagebox, scrolledtext
from datetime import datetime
from openpyxl import Workbook, load_workbook

# === Folder Setup ===
os.makedirs("known_faces", exist_ok=True)
os.makedirs("unknown_faces", exist_ok=True)

# === Excel Setup ===
excel_file = "student_registration.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Time", "Name", "Roll Number"])
    wb.save(excel_file)

# === GUI Setup ===
root = tk.Tk()
root.title("Student Face Recognition System")
root.geometry("560x520")
root.configure(bg="#f0f4f7")

title_label = tk.Label(root, text="Face Recognition Attendance System", font=("Helvetica", 18, "bold"), fg="#003366", bg="#f0f4f7")
title_label.pack(pady=20)

text_box = scrolledtext.ScrolledText(root, height=12, width=65, font=("Arial", 10))
text_box.pack(pady=10)

# === Register Student ===
def register_student():
    name = simpledialog.askstring("Input", "Enter Student Name:")
    roll = simpledialog.askstring("Input", "Enter Roll Number:")

    if not name or not roll:
        messagebox.showwarning("Missing Info", "Please enter both name and roll number.")
        return

    cam = cv2.VideoCapture(0)
    cv2.namedWindow("Capture Face - Press SPACE")

    captured = False
    while True:
        ret, frame = cam.read()
        if not ret:
            text_box.insert(tk.END, "Failed to access camera.\n")
            break

        cv2.imshow("Capture Face - Press SPACE", frame)
        key = cv2.waitKey(1)

        if key % 256 == 32:  # SPACE
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            face_locations = face_recognition.face_locations(rgb_frame)

            if face_locations:
                filename = f"{roll}_{name}.jpg"
                filepath = os.path.join("known_faces", filename)
                cv2.imwrite(filepath, frame)

                now = datetime.now()
                date_str = now.strftime("%Y-%m-%d")
                time_str = now.strftime("%H:%M:%S")

                wb = load_workbook(excel_file)
                ws = wb.active
                ws.append([date_str, time_str, name, roll])
                wb.save(excel_file)

                text_box.insert(tk.END, f"✔️ Registered: {name} ({roll}) on {date_str} at {time_str}\n")
                captured = True
            else:
                text_box.insert(tk.END, "⚠️ No face detected. Try again.\n")

        elif key % 256 == 27:  # ESC
            break

        if captured:
            break

    cam.release()
    cv2.destroyAllWindows()

# === Start Attendance with Unknown Face Handling ===
def start_attendance():
    known_encodings = []
    known_names = []
    known_rolls = []

    for filename in os.listdir("known_faces"):
        if filename.endswith(".jpg"):
            image = face_recognition.load_image_file(os.path.join("known_faces", filename))
            encoding = face_recognition.face_encodings(image)
            if encoding:
                known_encodings.append(encoding[0])
                roll, name = filename[:-4].split("_", 1)
                known_rolls.append(roll)
                known_names.append(name)

    if not known_encodings:
        messagebox.showinfo("No Students", "No registered faces found.")
        return

    cam = cv2.VideoCapture(0)
    cv2.namedWindow("Attendance - Press Q to quit")
    recognized = set()

    while True:
        ret, frame = cam.read()
        if not ret:
            text_box.insert(tk.END, "Failed to grab frame.\n")
            break

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb)
        face_encodings = face_recognition.face_encodings(rgb, face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=0.5)
            face_distances = face_recognition.face_distance(known_encodings, face_encoding)
            best_match_index = face_distances.argmin() if len(face_distances) > 0 else None

            if best_match_index is not None and matches[best_match_index]:
                name = known_names[best_match_index]
                roll = known_rolls[best_match_index]

                if roll not in recognized:
                    now = datetime.now()
                    date_str = now.strftime("%Y-%m-%d")
                    time_str = now.strftime("%H:%M:%S")

                    wb = load_workbook(excel_file)
                    ws = wb.active
                    ws.append([date_str, time_str, name, roll])
                    wb.save(excel_file)

                    text_box.insert(tk.END, f"🟢 Attendance marked: {name} ({roll}) at {time_str}\n")
                    recognized.add(roll)

                color = (0, 255, 0)
                label = f"{name} ({roll})"
            else:
                # Save unknown face
                now = datetime.now()
                timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
                unknown_filename = f"unknown_{timestamp}.jpg"
                cv2.imwrite(os.path.join("unknown_faces", unknown_filename), frame)
                text_box.insert(tk.END, f"⚠️ Unknown face detected at {timestamp}\n")

                color = (0, 0, 255)
                label = "Unknown"

            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.putText(frame, label, (left, top - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

        cv2.imshow("Attendance - Press Q to quit", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cam.release()
    cv2.destroyAllWindows()

# === Delete Student ===
def delete_student():
    roll = simpledialog.askstring("Delete", "Enter Roll Number to Delete:")
    if not roll:
        return

    deleted = False
    for file in os.listdir("known_faces"):
        if file.startswith(roll + "_"):
            os.remove(os.path.join("known_faces", file))
            text_box.insert(tk.END, f"🗑️ Deleted record and face for roll number: {roll}\n")
            deleted = True

    if not deleted:
        text_box.insert(tk.END, f"❌ No student found with roll number: {roll}\n")

# === Buttons ===
btn_style = {"font": ("Arial", 12), "width": 20, "height": 2}

tk.Button(root, text="➕ Register Student", bg="#4CAF50", fg="white", command=register_student, **btn_style).pack(pady=5)
tk.Button(root, text="🟢 Start Attendance", bg="#2196F3", fg="white", command=start_attendance, **btn_style).pack(pady=5)
tk.Button(root, text="🗑️ Delete Student", bg="#f44336", fg="white", command=delete_student, **btn_style).pack(pady=5)
tk.Button(root, text="❌ Exit", bg="#9E9E9E", fg="white", command=root.destroy, **btn_style).pack(pady=5)

# === Run ===
root.mainloop()
